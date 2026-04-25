from pathlib import Path
from openpyxl import load_workbook
from django.conf import settings

from apps.core.exceptions import RepositoryError


class ExcelWorkbookManager:
    """
    Excelファイルの読み込み・保存を担当する共通クラス。
    ViewやServiceから直接Excelを触らないようにする。
    """

    def __init__(self, filepath: str | None = None):
        self.filepath = filepath or settings.PMI_EXCEL_PATH

    def load(self):
        path = Path(self.filepath)

        if not path.exists():
            raise RepositoryError(
                f"Excel file not found: {path}\n"
                "data/master_data.xlsx を配置してください。"
            )

        return load_workbook(path)

    def save(self, workbook) -> None:
        workbook.save(self.filepath)

    def get_sheet(self, workbook, sheet_name: str):
        if sheet_name not in workbook.sheetnames:
            raise RepositoryError(
                f"Sheet not found: {sheet_name}\n"
                f"Excel内に '{sheet_name}' シートを作成してください。"
            )

        return workbook[sheet_name]